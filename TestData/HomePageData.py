import openpyxl


class HomePageData:
    # test_HomePage_data = [{"email": "fake@gmail.com", "password": "aloha123", "gender": "Female"},
                          # {"email": "4fake4@gmail.com", "password": "aloha123!", "gender": "Male"}]

    # if you want to call the method directly like HomePageData.get_test_data, w\o crass instance creation, make this class static
    # we do not need self parameter for static methods
    @staticmethod
    def get_test_data(test_case_name):

        book = openpyxl.load_workbook("C:\\DATA\\QA\\Python\\python.xlsx")
        sheet = book.active
        dictionary = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dictionary]  # we should send dictionary to a list
