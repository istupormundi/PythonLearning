import openpyxl

book = openpyxl.load_workbook("C:\\DATA\\QA\\Python\\python.xlsx")
sheet = book.active
dictionary = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "ALOHA"
print(sheet.cell(row=2, column=2).value)

print("max row: " + str(sheet.max_row))
print("max col: " + str(sheet.max_column))

print(sheet['A5'].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        print(i)
        for j in range(2, sheet.max_column + 1):
            # print("(i,j) = " "(" + str(i) + ")" + "(" + str(j) + ")")
            dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            # print(sheet.cell(row=i, column=j).value)
print(dictionary)

